import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Dashboard de Disponibilidade", page_icon="🏭", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: #0f1117; color: #e2e8f0; }
[data-testid="stSidebar"] { background: #161b27 !important; border-right: 1px solid #1e2d45; }
[data-testid="stSidebar"] * { color: #cbd5e1 !important; }
[data-testid="stSidebar"] h1,[data-testid="stSidebar"] h2,[data-testid="stSidebar"] h3 { color: #7dd3fc !important; }
h1 { color: #f1f5f9 !important; font-weight: 700; letter-spacing: -0.5px; }
h2, h3 { color: #cbd5e1 !important; font-weight: 600; }
.kpi-grid { display:flex; gap:14px; flex-wrap:wrap; margin:16px 0 24px 0; }
.kpi-card {
    flex:1 1 160px; background:#161b27; border:1px solid #1e2d45;
    border-radius:14px; padding:18px 20px; position:relative; overflow:hidden;
}
.kpi-card::before { content:''; position:absolute; top:0; left:0; right:0; height:3px; }
.kpi-card.green::before  { background:linear-gradient(90deg,#22c55e,#86efac); }
.kpi-card.orange::before { background:linear-gradient(90deg,#f59e0b,#fde68a); }
.kpi-card.red::before    { background:linear-gradient(90deg,#ef4444,#fca5a5); }
.kpi-card.blue::before   { background:linear-gradient(90deg,#3b82f6,#93c5fd); }
.kpi-card.purple::before { background:linear-gradient(90deg,#8b5cf6,#c4b5fd); }
.kpi-label { font-size:.7rem; font-weight:600; text-transform:uppercase; letter-spacing:1.2px; color:#64748b; margin-bottom:8px; }
.kpi-value { font-size:2rem; font-weight:700; color:#f1f5f9; font-family:'DM Mono',monospace; line-height:1; }
.kpi-sub   { font-size:.72rem; color:#475569; margin-top:6px; }
.stTabs [data-baseweb="tab-list"] { gap:4px; background:#161b27; border-radius:10px; padding:4px; border:1px solid #1e2d45; }
.stTabs [data-baseweb="tab"] { border-radius:7px; padding:8px 18px; font-size:.82rem; font-weight:500; color:#64748b !important; background:transparent; border:none; }
.stTabs [aria-selected="true"] { background:#1e3a5f !important; color:#7dd3fc !important; }
hr { border-color:#1e2d45; margin:20px 0; }
[data-testid="stDataFrame"] { border-radius:10px; overflow:hidden; }
.section-header { font-size:.75rem; font-weight:600; text-transform:uppercase; letter-spacing:2px; color:#3b82f6; margin:28px 0 12px 0; border-bottom:1px solid #1e2d45; padding-bottom:8px; }
.stDownloadButton button { background:linear-gradient(135deg,#1e3a5f,#2563eb) !important; color:white !important; border:none !important; border-radius:8px !important; font-weight:600 !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
#  MAPEAMENTO POR ID DE TRABALHO  (fonte de verdade)
#  Edite aqui para adicionar/mover IDs entre categorias
# ─────────────────────────────────────────────────────────────
MAPEAMENTO_ATIVIDADES: dict[str, str] = {
    # ── Tempo Indireto Produtivo (conta no denominador como produtivo) ──
    "OPTR00039319": "Tempo Indireto Produtivo",   # KAIZEN
    "OPTR00039324": "Tempo Indireto Produtivo",   # Treinamentos
    "OPTR00039364": "Tempo Indireto Produtivo",   # Apoio
    "OPTR00039315": "Tempo Indireto Produtivo",   # Auditorias
    "OPTR00039323": "Tempo Indireto Produtivo",   # Eventos GG
    "OPTR00039318": "Tempo Indireto Produtivo",   # Gerenciamento de risco
    "OPTR00039320": "Tempo Indireto Produtivo",   # Laboral
    "OPTR00039536": "Tempo Indireto Produtivo",   # FCPIE
    "OPTR00039317": "Tempo Indireto Produtivo",   # DDP
    # ── Paradas Planejadas ──
    "OPTR00039530": "Paradas Planejadas",          # Almoço
    "OPTR00039322": "Paradas Planejadas",          # Café e Pausas
    "OPTR00039314": "Paradas Planejadas",          # Setup / Preparação
    "OPTR00039310": "Paradas Planejadas",          # Manutenção preventiva
    "OPTR00039312": "Paradas Planejadas",          # Retrabalho (intervalo)
    "OPTR00044938": "Paradas Planejadas",          # Retrabalho MP/ME
    # ── Paradas Não Planejadas ──
    "OPTR00036729": "Paradas Não Planejadas",      # Aguardando trabalho
    "OPTR00042380": "Paradas Não Planejadas",      # Aguardando/Esperando alocação
    "OPTR00039308": "Paradas Não Planejadas",      # Aguardando/Esperando
    "OPTR00039359": "Paradas Não Planejadas",      # Aguardando inspeção
    "OPTR00039309": "Paradas Não Planejadas",      # Falhas / Quebras
    "OPTR00039311": "Paradas Não Planejadas",      # Qualidade
    "OPTR00039313": "Paradas Não Planejadas",      # Sem recursos
}

DIAS_SEMANA_PT = {0:"Segunda",1:"Terça",2:"Quarta",3:"Quinta",4:"Sexta",5:"Sábado",6:"Domingo"}


# ─────────────────────────────────────────────────────────────
#  UTILITÁRIOS
# ─────────────────────────────────────────────────────────────

def hhmm(h: float) -> str:
    """Decimal → 'HH:MM'"""
    if pd.isna(h) or h < 0: return "00:00"
    hi = int(h); m = round((h - hi) * 60)
    if m == 60: hi += 1; m = 0
    return f"{hi:02d}:{m:02d}"


def jornada_do_dia(data, horas_padrao: float, reducao_sexta: float) -> float:
    return horas_padrao - reducao_sexta if pd.Timestamp(data).dayofweek == 4 else horas_padrao


def categorizar(row, mapeamento: dict) -> str:
    tipo  = str(row.get("Tipo de registro do diário", "")).strip()
    ident = str(row.get("Ident. do trabalho", "")).strip()
    if tipo in ("Registro de entrada", "Registro de saída"): return "Outros"
    if tipo == "Processo":    return "Tempo Produtivo"
    if tipo == "Interrupção": return "Paradas Não Planejadas"
    return mapeamento.get(ident, "Outros")


def calcular_horas_exclusivas(sub: pd.DataFrame) -> dict:
    """
    Soma as horas resolvendo sobreposições cruzadas (Sweep-line algorithm).
    - Mesma categoria: funde os tempos (trata 2 OPs como 1 só).
    - Categorias diferentes: a que tem menor número em 'PRIORIDADE' ganha o tempo.
    """
    # 1. Defina a hierarquia de quem rouba o tempo de quem
    PRIORIDADE = {
        "Paradas Não Planejadas": 1,  # Máquina quebrou? Interrompe tudo.
        "Paradas Planejadas": 2,      # Intervalo/Almoço vence o produtivo.
        "Tempo Produtivo": 3,         # Se não está parado, está produzindo.
        "Tempo Indireto Produtivo": 4,# Apoios e 5S ficam no fim da fila.
        "Outros": 5
    }

    df_valid = sub.dropna(subset=["Inicio_ts", "Fim_ts"])
    if df_valid.empty:
        return {cat: 0.0 for cat in PRIORIDADE.keys()}

    # 2. Criar eventos de início (1) e fim (-1)
    eventos = []
    for _, r in df_valid.iterrows():
        if r["Inicio_ts"] < r["Fim_ts"]:
            eventos.append((r["Inicio_ts"], 1, r["Categoria"]))
            eventos.append((r["Fim_ts"], -1, r["Categoria"]))

    # Ordena cronologicamente. Desempate: Inícios antes de Fins.
    eventos.sort(key=lambda x: (x[0], -x[1]))

    tempos = {c: 0.0 for c in PRIORIDADE.keys()}
    ativas = {}
    ultimo_tempo = None

    # 3. Varredura da linha do tempo
    for tempo, acao, cat in eventos:
        if ultimo_tempo is not None and tempo > ultimo_tempo:
            # Descobre quem está rodando e quem ganha
            cats_ativas = [c for c, count in ativas.items() if count > 0]
            if cats_ativas:
                vencedora = min(cats_ativas, key=lambda c: PRIORIDADE.get(c, 99))
                duracao = (tempo - ultimo_tempo).total_seconds() / 3600.0
                tempos[vencedora] += duracao

        # Atualiza o status da categoria atual (abriu ou fechou)
        ativas[cat] = ativas.get(cat, 0) + acao
        ultimo_tempo = tempo

    return tempos


# ─────────────────────────────────────────────────────────────
#  PROCESSAMENTO
# ─────────────────────────────────────────────────────────────

def processar(df_raw: pd.DataFrame, horas_padrao: float, reducao_sexta: float,
              mapeamento: dict) -> pd.DataFrame:
    df = df_raw.copy()
    df = df.drop_duplicates()
    df.columns = df.columns.str.strip()

    df["Data do perfil"] = pd.to_datetime(df["Data do perfil"], dayfirst=True, errors="coerce")

    for col in ["Hora inicial", "Hora final"]:
        df[col] = pd.to_datetime(df[col], errors="coerce").dt.time

    df.dropna(subset=["Hora inicial", "Hora final", "Data do perfil"], inplace=True)

    df["Inicio_ts"] = pd.to_datetime(df["Data do perfil"].astype(str) + " " + df["Hora inicial"].astype(str))
    df["Fim_ts"]    = pd.to_datetime(df["Data do perfil"].astype(str) + " " + df["Hora final"].astype(str))
    df.loc[df["Fim_ts"] < df["Inicio_ts"], "Fim_ts"] += pd.Timedelta(days=1)
    df["Duracao_Horas"] = (df["Fim_ts"] - df["Inicio_ts"]).dt.total_seconds() / 3600

    df["Categoria"] = df["Ident. do trabalho"].map(mapeamento).fillna("Outros")

    df.loc[df["Tipo de registro do diário"] == "Processo", "Categoria"] = "Tempo Produtivo"
    df.loc[df["Tipo de registro do diário"] == "Interrupção", "Categoria"] = "Paradas Não Planejadas"
    df.loc[df["Tipo de registro do diário"] == "Atividade indireta", "Categoria"] = \
        df["Ident. do trabalho"].map(mapeamento).fillna("Tempo Indireto Produtivo")
    df.loc[df["Tipo de registro do diário"].isin(["Registro de entrada","Registro de saída"]), "Categoria"] = "Outros"
    df["Data"]          = df["Data do perfil"].dt.date
    df["Dia_Semana"]    = df["Data do perfil"].dt.dayofweek.map(DIAS_SEMANA_PT)
    df["Jornada_Dia"]   = df["Data"].apply(lambda d: jornada_do_dia(d, horas_padrao, reducao_sexta))
    df["Duracao_HH:MM"] = df["Duracao_Horas"].apply(hhmm)

    for col in ["Referência", "N° oper.", "Ident. do trabalho", "Descrição"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].astype(str)

    return df


def agregar_pessoa(df: pd.DataFrame, horas_padrao: float) -> pd.DataFrame:
    CATS = ["Tempo Produtivo", "Paradas Planejadas", "Paradas Não Planejadas", "Tempo Indireto Produtivo"]
    rows = []
    for nome, grp in df.groupby("Nome"):
        r = {"Nome": nome}
        
        # APLICA A NOVA LÓGICA DE PRIORIDADE E FUSÃO
        tempos = calcular_horas_exclusivas(grp)
        for cat in CATS:
            r[cat] = tempos.get(cat, 0.0)
            
        # Jornada = soma dos dias únicos com jornada própria (sexta reduzida)
        r["Jornada_Total"] = grp.groupby("Data")["Jornada_Dia"].first().sum()
        
        # Disponibilidade = TP / (Jornada Total - Paradas Planejadas)
        base = r["Jornada_Total"] - r["Paradas Planejadas"]
        r["Disponibilidade (%)"] = round((r["Tempo Produtivo"] / base) * 100, 2) if base > 0 else 0.0
        
        rows.append(r)
        
    pivot = pd.DataFrame(rows)
    for cat in CATS:
        if cat not in pivot.columns: pivot[cat] = 0.0
        pivot[cat + "_HH:MM"] = pivot[cat].apply(hhmm)
        
    pivot["Jornada_Total_HH:MM"] = pivot["Jornada_Total"].apply(hhmm)
    return pivot


def agregar_data(df: pd.DataFrame, horas_padrao: float, reducao_sexta: float) -> pd.DataFrame:
    CATS = ["Tempo Produtivo", "Paradas Planejadas", "Paradas Não Planejadas"]
    rows = []
    for data, grp_data in df.groupby("Data"):
        r = {"Data": data}
        r["Dia_Semana"]  = DIAS_SEMANA_PT[pd.Timestamp(data).dayofweek]
        r["Jornada_Dia"] = jornada_do_dia(data, horas_padrao, reducao_sexta)
        r["Num_Pessoas"] = grp_data["Nome"].nunique()

        # APLICA A NOVA LÓGICA DE PRIORIDADE E FUSÃO (separado por pessoa)
        tempos_totais_dia = {c: 0.0 for c in CATS}
        for nome, grp_pessoa in grp_data.groupby("Nome"):
            tempos_pessoa = calcular_horas_exclusivas(grp_pessoa)
            for cat in CATS:
                tempos_totais_dia[cat] += tempos_pessoa.get(cat, 0.0)

        for cat in CATS:
            r[cat] = tempos_totais_dia[cat]

        # Base = (Jornada do Dia * Número de Pessoas) - Paradas Planejadas
        jornada_total_dia = r["Jornada_Dia"] * r["Num_Pessoas"]
        base = jornada_total_dia - r["Paradas Planejadas"]
        
        r["Disponibilidade (%)"] = round((r["Tempo Produtivo"] / base) * 100, 2) if base > 0 else 0.0
        rows.append(r)

    pivot = pd.DataFrame(rows)
    for cat in CATS:
        pivot[cat + "_HH:MM"] = pivot[cat].apply(hhmm)
    return pivot

def _distribuir_horas_op(df_proc: pd.DataFrame) -> dict:
    """
    Para registros de Tempo Produtivo de uma (pessoa, dia),
    divide a linha do tempo em micro-slots onde o conjunto de operações ativas muda.
    Cada slot tem sua duração real dividida proporcionalmente entre as operações ativas.
    Retorna {(ref, oper, desc, ident): horas_reais}
    """
    if df_proc.empty:
        return {}

    col_oper  = "N° oper."           if "N° oper."           in df_proc.columns else None
    col_desc  = "Descrição"          if "Descrição"          in df_proc.columns else None
    col_ident = "Ident. do trabalho" if "Ident. do trabalho" in df_proc.columns else None

    def chave(r):
        return (r["Referência"],
                r[col_oper]  if col_oper  else "",
                r[col_desc]  if col_desc  else "",
                r[col_ident] if col_ident else "")

    # Pontos de mudança na linha do tempo
    eventos = sorted({ts for _, r in df_proc.iterrows() for ts in (r["Inicio_ts"], r["Fim_ts"])})

    acumulado: dict = {}
    for i in range(len(eventos) - 1):
        t_ini, t_fim = eventos[i], eventos[i + 1]
        if t_ini >= t_fim:
            continue
        dur_slot = (t_fim - t_ini).total_seconds() / 3600
        ativas = df_proc[(df_proc["Inicio_ts"] <= t_ini) & (df_proc["Fim_ts"] >= t_fim)]
        if ativas.empty:
            continue
        parcela = dur_slot / len(ativas)
        for _, r in ativas.iterrows():
            k = chave(r)
            acumulado[k] = acumulado.get(k, 0.0) + parcela
    return acumulado


def agregar_op(df: pd.DataFrame):
    df_p = df[df["Categoria"] == "Tempo Produtivo"].copy()
    if df_p.empty:
        return pd.DataFrame(), pd.DataFrame()

    col_oper  = "N° oper."           if "N° oper."           in df_p.columns else None
    col_desc  = "Descrição"          if "Descrição"          in df_p.columns else None
    col_ident = "Ident. do trabalho" if "Ident. do trabalho" in df_p.columns else None

    from collections import defaultdict
    horas_oe  = defaultdict(float)   # {(ref,oper,desc,ident): h}
    pessoas_r = defaultdict(set)     # {ref: {nomes}}
    dias_r    = defaultdict(set)     # {ref: {datas}}
    regs_r    = defaultdict(int)     # {ref: n registros brutos}

    for (nome, data), grp in df_p.groupby(["Nome", "Data"]):
        distrib = _distribuir_horas_op(grp)
        for (ref, oper, desc, ident), h in distrib.items():
            horas_oe[(ref, oper, desc, ident)] += h
            pessoas_r[ref].add(nome)
            dias_r[ref].add(data)
        for ref in grp["Referência"].unique():
            regs_r[ref] += len(grp[grp["Referência"] == ref])

    # ── Por OP ──
    op_rows = []
    for ref in sorted(pessoas_r.keys()):
        h = sum(v for (r, *_), v in horas_oe.items() if r == ref)
        op_rows.append({"Ordem de Produção": ref, "Horas": round(h, 4),
                        "Registros": regs_r[ref],
                        "Nº Pessoas": len(pessoas_r[ref]),
                        "Nº Dias": len(dias_r[ref])})
    op = pd.DataFrame(op_rows)
    op["Horas_HH:MM"] = op["Horas"].apply(hhmm)
    op = op.sort_values("Horas", ascending=False).reset_index(drop=True)

    # ── Por OP + Operação ──
    op_oper_rows = []
    for (ref, oper, desc, ident), h in horas_oe.items():
        row = {"Ordem de Produção": ref, "Horas": round(h, 4)}
        if col_oper:  row["Operação"]        = oper
        if col_desc:  row["Descrição"]       = desc
        if col_ident: row["Ident. Trabalho"] = ident
        op_oper_rows.append(row)
    op_oper = pd.DataFrame(op_oper_rows)
    op_oper["Horas_HH:MM"] = op_oper["Horas"].apply(hhmm)
    op_oper = op_oper.sort_values("Horas", ascending=False).reset_index(drop=True)

    return op, op_oper



def agregar_paradas(df: pd.DataFrame) -> pd.DataFrame:
    sub = df[df["Categoria"].isin(["Paradas Planejadas","Paradas Não Planejadas"])].copy()
    if sub.empty: return pd.DataFrame(columns=["Descrição","Tipo","Horas","Horas_HH:MM"])
    p = sub.groupby(["Descrição","Categoria"])["Duracao_Horas"].sum().reset_index()
    p.columns = ["Descrição","Tipo","Horas"]
    p["Horas_HH:MM"] = p["Horas"].apply(hhmm)
    return p.sort_values("Horas", ascending=False)


def detectar_sobreposicoes(df: pd.DataFrame) -> list:
    cats = ["Tempo Produtivo","Paradas Planejadas","Paradas Não Planejadas"]
    out = []
    for (nome, data), grp in df[df["Categoria"].isin(cats)].groupby(["Nome","Data"]):
        for cat in cats:
            sub = grp[grp["Categoria"]==cat].sort_values("Inicio_ts").reset_index(drop=True)
            for i in range(len(sub)-1):
                if sub.loc[i+1,"Inicio_ts"] < sub.loc[i,"Fim_ts"]:
                    out.append({"Pessoa":nome,"Data":data,"Categoria":cat,
                                "Reg. A":sub.loc[i,"Descrição"],"Reg. B":sub.loc[i+1,"Descrição"]})
    return out


# ─────────────────────────────────────────────────────────────
#  EXCEL
# ─────────────────────────────────────────────────────────────

def gerar_excel(tp, td, par, op, op_oper, horas_padrao):
    wb = Workbook()
    hf  = PatternFill("solid", fgColor="1E3A5F")
    hft = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ttf = Font(bold=True, color="1E3A5F", name="Calibri", size=12)
    af  = PatternFill("solid", fgColor="EAF2FB")
    brd = Border(**{s: Side(style="thin", color="CCCCCC") for s in ("left","right","top","bottom")})
    ctr = Alignment(horizontal="center", vertical="center")

    def sheet(ws, df, sr, title):
        ws.cell(sr, 1, title).font = ttf
        hr = sr + 1
        for j, c in enumerate(df.columns, 1):
            cell = ws.cell(hr, j, c); cell.fill=hf; cell.font=hft; cell.alignment=ctr; cell.border=brd
        for i, row in enumerate(df.itertuples(index=False), hr+1):
            for j, v in enumerate(row, 1):
                c = ws.cell(i, j, v); c.border=brd; c.alignment=ctr
                if i%2==0: c.fill=af
                if isinstance(v, float): c.number_format="0.00"
        for j, col in enumerate(df.columns, 1):
            max_len = df[col].astype(str).map(len).max() if not df.empty else len(col)  
            w = max(max_len, len(col)) + 4  
            ws.column_dimensions[get_column_letter(j)].width = min(w, 45)

    ws1 = wb.active; ws1.title="Resumo"; ws1.sheet_view.showGridLines=False
    ws1["A1"] = f"Dashboard de Disponibilidade  |  Jornada: {horas_padrao}h"
    ws1["A1"].font=Font(bold=True,color="FFFFFF",name="Calibri",size=13)
    ws1["A1"].fill=PatternFill("solid",fgColor="1E3A5F")
    ws1["A1"].alignment=ctr; ws1.merge_cells("A1:G1"); ws1.row_dimensions[1].height=30

    cols_p = ["Nome","Tempo Produtivo_HH:MM","Paradas Planejadas_HH:MM",
              "Paradas Não Planejadas_HH:MM","Tempo Indireto Produtivo_HH:MM",
              "Jornada_Total_HH:MM","Disponibilidade (%)"]
    sheet(ws1, tp[[c for c in cols_p if c in tp.columns]], 3, "Disponibilidade por Pessoa")

    ws2 = wb.create_sheet("Por Data"); ws2.sheet_view.showGridLines=False
    cols_d = ["Data","Dia_Semana","Jornada_Dia","Num_Pessoas",
              "Tempo Produtivo_HH:MM","Paradas Planejadas_HH:MM",
              "Paradas Não Planejadas_HH:MM","Disponibilidade (%)"]
    sheet(ws2, td[[c for c in cols_d if c in td.columns]], 2, "Por Data")

    ws3 = wb.create_sheet("Paradas"); ws3.sheet_view.showGridLines=False
    if not par.empty: sheet(ws3, par, 2, "Análise de Paradas")

    ws4 = wb.create_sheet("Horas por OP"); ws4.sheet_view.showGridLines=False
    if not op.empty: sheet(ws4, op, 2, "Horas por OP")

    if not op_oper.empty:
        ws5 = wb.create_sheet("OP + Operações"); ws5.sheet_view.showGridLines=False
        sheet(ws5, op_oper, 2, "Horas por OP e Operação")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## ⚙️ Configurações")
    st.markdown("---")
    st.markdown("**Jornada de Trabalho**")
    horas_padrao = st.number_input("Horas/dia (padrão)", 1.0, 24.0, 10.0, 0.5)
    reducao_sexta = st.number_input("Redução Sexta-feira (h)", 0.0, 4.0, 1.0, 0.5,
        help="Sexta = horas_padrao − esta redução")
    st.caption(f"📅 Seg–Qui: **{horas_padrao:.1f}h** | Sex: **{horas_padrao-reducao_sexta:.1f}h**")
    st.markdown("---")
    arquivo = st.file_uploader("📂 Planilha (.xlsx)", type=["xlsx"])
    st.markdown("---")

    # Editor de mapeamento dinâmico
    with st.expander("🗂️ Mapeamento de IDs (avançado)", expanded=False):
        st.caption("Adicione IDs extras no formato `OPTR00012345=Paradas Não Planejadas`")
        extra_ids = st.text_area("Um por linha:", height=120,
            placeholder="OPTR00099999=Paradas Não Planejadas\nOPTR00088888=Paradas Planejadas")
        if extra_ids.strip():
            for linha in extra_ids.strip().split("\n"):
                if "=" in linha:
                    k, v = linha.split("=", 1)
                    MAPEAMENTO_ATIVIDADES[k.strip()] = v.strip()

    st.markdown("**Legenda**")
    st.markdown("🟢 **Produtivo** — Tipo: Processo")
    st.markdown("🔵 **T.Ind.Prod.** — Kaizen/Treino/Laboral...")
    st.markdown("🟡 **Plan.** — Café/Almoço/Setup/Retrabalho...")
    st.markdown("🔴 **N.Plan.** — Aguardando/Esperando/Falhas...")


# ─────────────────────────────────────────────────────────────
#  HEADER
# ─────────────────────────────────────────────────────────────

st.markdown("""
<div style="display:flex;align-items:center;gap:14px;margin-bottom:4px;">
  <div style="font-size:2.4rem;">🏭</div>
  <div>
    <h1 style="margin:0;font-size:1.7rem;">Dashboard de Disponibilidade</h1>
    <p style="margin:0;color:#475569;font-size:.85rem;">Análise de produção · Dynamics 365 F&amp;O</p>
  </div>
</div>
""", unsafe_allow_html=True)
st.markdown("<hr>", unsafe_allow_html=True)

PL = dict(plot_bgcolor="#161b27", paper_bgcolor="#0f1117",
          font=dict(color="#94a3b8", family="DM Sans"),
          margin=dict(l=10, r=10, t=40, b=10))


# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────

if arquivo:
    try:
        df_raw = pd.read_excel(arquivo)
        df = processar(df_raw, horas_padrao, reducao_sexta, MAPEAMENTO_ATIVIDADES)

        # ── Filtros ──
        datas_all = sorted(df["Data"].unique())
        cf1, cf2, cf3 = st.columns([2, 2, 2])
        with cf1:
            modo = st.selectbox("📅 Período", ["Todos os dias","Selecionar dias","Mês específico"])
        with cf2:
            datas_sel, mes_sel = None, None
            if modo == "Selecionar dias":
                datas_sel = st.multiselect("Dias", datas_all, default=datas_all[:min(5,len(datas_all))])
            elif modo == "Mês específico":
                meses = sorted({(pd.Timestamp(d).year, pd.Timestamp(d).month) for d in datas_all})
                ml = {f"{m[0]}-{m[1]:02d}": m for m in meses}
                mes_sel = ml[st.selectbox("Mês", list(ml.keys()))]
            else:
                st.markdown(" ")
        with cf3:
            pessoas_all = sorted(df["Nome"].unique())
            pessoas_sel = st.multiselect("👤 Pessoas", pessoas_all, default=pessoas_all)

        df_f = df.copy()
        if modo == "Selecionar dias" and datas_sel:
            df_f = df_f[df_f["Data"].isin(datas_sel)]
        elif modo == "Mês específico" and mes_sel:
            df_f = df_f[(pd.to_datetime(df_f["Data"]).dt.year == mes_sel[0]) &
                        (pd.to_datetime(df_f["Data"]).dt.month == mes_sel[1])]
        if pessoas_sel:
            df_f = df_f[df_f["Nome"].isin(pessoas_sel)]

        if df_f.empty:
            st.warning("Nenhum dado para os filtros selecionados."); st.stop()

        tp_pessoa = agregar_pessoa(df_f, horas_padrao)
        tp_data   = agregar_data(df_f, horas_padrao, reducao_sexta)
        op_tot, op_oper = agregar_op(df_f)
        paradas   = agregar_paradas(df_f)

        # ── KPIs ──
        tp_h  = tp_pessoa["Tempo Produtivo"].sum()
        pp_h  = tp_pessoa["Paradas Planejadas"].sum()
        pnp_h = tp_pessoa["Paradas Não Planejadas"].sum()
        tip_h = tp_pessoa["Tempo Indireto Produtivo"].sum() if "Tempo Indireto Produtivo" in tp_pessoa.columns else 0.0
        n_pes = len(tp_pessoa)
        n_dia = df_f["Data"].nunique()
        
        # NOVA FÓRMULA PARA O KPI GERAL
        jornada_geral = tp_pessoa["Jornada_Total"].sum()
        base_geral = jornada_geral - pp_h
        disp_geral = round((tp_h / base_geral) * 100, 1) if base_geral > 0 else 0.0
        cor = "green" if disp_geral >= 85 else "orange" if disp_geral >= 70 else "red"

        st.markdown(f"""
        <div class="kpi-grid">
          <div class="kpi-card {cor}">
            <div class="kpi-label">Disponibilidade Geral</div>
            <div class="kpi-value">{disp_geral:.1f}%</div>
            <div class="kpi-sub">TP ÷ (TP + Paradas N.Plan.)</div>
          </div>
          <div class="kpi-card blue">
            <div class="kpi-label">Pessoas · Dias</div>
            <div class="kpi-value">{n_pes}</div>
            <div class="kpi-sub">{n_dia} dia(s) analisado(s)</div>
          </div>
          <div class="kpi-card green">
            <div class="kpi-label">Tempo Produtivo</div>
            <div class="kpi-value">{hhmm(tp_h)}</div>
            <div class="kpi-sub">{tp_h:.2f} h</div>
          </div>
          <div class="kpi-card orange">
            <div class="kpi-label">Paradas Planejadas</div>
            <div class="kpi-value">{hhmm(pp_h)}</div>
            <div class="kpi-sub">{pp_h:.2f} h</div>
          </div>
          <div class="kpi-card red">
            <div class="kpi-label">Paradas N.Planejadas</div>
            <div class="kpi-value">{hhmm(pnp_h)}</div>
            <div class="kpi-sub">{pnp_h:.2f} h</div>
          </div>
          <div class="kpi-card purple">
            <div class="kpi-label">T.Ind. Produtivo</div>
            <div class="kpi-value">{hhmm(tip_h)}</div>
            <div class="kpi-sub">Kaizen · Treino · Laboral</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Tabs ──
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📊 Disponibilidade", "⏸️ Paradas",
            "🏭 Ordens de Produção", "📅 Por Data", "📋 Dados Brutos"
        ])

        # ── TAB 1 ──
        with tab1:
            c1, c2 = st.columns([3, 2])
            with c1:
                cores_bar = ["#22c55e" if v>=85 else "#f59e0b" if v>=70 else "#ef4444"
                             for v in tp_pessoa["Disponibilidade (%)"]]
                fig = go.Figure(go.Bar(
                    x=tp_pessoa["Nome"], y=tp_pessoa["Disponibilidade (%)"],
                    marker_color=cores_bar,
                    text=[f"{v:.1f}%" for v in tp_pessoa["Disponibilidade (%)"]],
                    textposition="outside", textfont=dict(color="#e2e8f0", size=11)
                ))
                fig.add_hline(y=85, line_dash="dash", line_color="#22c55e", line_width=1.5,
                              annotation_text="Meta 85%", annotation_font_color="#22c55e")
                fig.update_layout(title=dict(text="Disponibilidade por Pessoa (%)", font=dict(color="#e2e8f0")),
                    xaxis=dict(tickangle=-30, gridcolor="#1e2d45", color="#64748b"),
                    yaxis=dict(range=[0,100], gridcolor="#1e2d45", color="#64748b"), height=420, **PL)
                st.plotly_chart(fig, use_container_width=True)

            with c2:
                fig2 = go.Figure(go.Pie(
                    labels=["Produtivo","Par. Planejadas","Par. N.Planejadas","T.Ind.Prod."],
                    values=[tp_h, pp_h, pnp_h, tip_h],
                    marker_colors=["#22c55e","#f59e0b","#ef4444","#8b5cf6"],
                    hole=0.55, textinfo="label+percent",
                    textfont=dict(color="#e2e8f0", size=10)
                ))
                fig2.update_layout(title=dict(text="Composição da Jornada", font=dict(color="#e2e8f0")),
                    height=420, showlegend=False, **PL)
                st.plotly_chart(fig2, use_container_width=True)

            dcols = ["Nome","Tempo Produtivo_HH:MM","Paradas Planejadas_HH:MM",
                     "Paradas Não Planejadas_HH:MM","Tempo Indireto Produtivo_HH:MM",
                     "Jornada_Total_HH:MM","Disponibilidade (%)"]
            dcols = [c for c in dcols if c in tp_pessoa.columns]
            df_show = tp_pessoa[dcols].rename(columns={
                "Tempo Produtivo_HH:MM":"Produtivo",
                "Paradas Planejadas_HH:MM":"Par. Plan.",
                "Paradas Não Planejadas_HH:MM":"Par. N.Plan.",
                "Tempo Indireto Produtivo_HH:MM":"T.Ind.Prod.",
                "Jornada_Total_HH:MM":"Jornada",
                "Disponibilidade (%)":"Disp. (%)"
            })
            st.dataframe(df_show.style
                .background_gradient(subset=["Disp. (%)"], cmap="RdYlGn", vmin=0, vmax=100)
                .format({"Disp. (%)":"{:.2f}%"}),
                use_container_width=True, hide_index=True)

        # ── TAB 2 ──
        with tab2:
            if paradas.empty:
                st.info("Sem paradas registradas.")
            else:
                c1, c2 = st.columns(2)
                with c1:
                    fp = px.bar(paradas.head(15), x="Horas", y="Descrição", color="Tipo",
                        orientation="h", text="Horas_HH:MM", title="Top 15 Causas de Parada",
                        color_discrete_map={"Paradas Planejadas":"#f59e0b","Paradas Não Planejadas":"#ef4444"})
                    fp.update_layout(height=450, **PL,
                        xaxis=dict(gridcolor="#1e2d45",color="#64748b"),
                        yaxis=dict(gridcolor="#1e2d45",color="#64748b"),
                        title=dict(font=dict(color="#e2e8f0")),
                        legend=dict(font=dict(color="#94a3b8"),bgcolor="#161b27"))
                    st.plotly_chart(fp, use_container_width=True)
                with c2:
                    fp2 = go.Figure(go.Pie(labels=paradas["Descrição"], values=paradas["Horas"],
                        hole=0.4, textinfo="label+percent", textfont=dict(color="#e2e8f0",size=10)))
                    fp2.update_layout(title=dict(text="Distribuição de Paradas",font=dict(color="#e2e8f0")),
                        height=450, showlegend=False, **PL)
                    st.plotly_chart(fp2, use_container_width=True)
                st.dataframe(paradas[["Tipo","Descrição","Horas_HH:MM","Horas"]].rename(
                    columns={"Horas_HH:MM":"Duração","Horas":"h (dec.)"}).style.format({"h (dec.)":"{:.2f}"}),
                    use_container_width=True, hide_index=True)

        # ── TAB 3 ──
        with tab3:
            if op_tot.empty:
                st.info("Nenhuma OP encontrada.")
            else:
                st.markdown('<div class="section-header">Horas por Ordem de Produção</div>', unsafe_allow_html=True)
                c1, c2 = st.columns([3, 2])
                with c1:
                    top = op_tot.head(20)
                    fo = go.Figure(go.Bar(x=top["Ordem de Produção"], y=top["Horas"],
                        marker=dict(color=top["Horas"],colorscale="Blues",showscale=False),
                        text=top["Horas_HH:MM"], textposition="outside",
                        textfont=dict(color="#e2e8f0",size=10)))
                    fo.update_layout(title=dict(text="Top 20 OPs",font=dict(color="#e2e8f0")),
                        xaxis=dict(tickangle=-30,gridcolor="#1e2d45",color="#64748b"),
                        yaxis=dict(gridcolor="#1e2d45",color="#64748b",title="Horas"),
                        height=420, **PL)
                    st.plotly_chart(fo, use_container_width=True)
                with c2:
                    ft = px.treemap(op_tot.head(20), path=["Ordem de Produção"], values="Horas",
                        title="Mapa de OPs", color="Horas", color_continuous_scale="Blues")
                    ft.update_layout(height=420, **PL, title=dict(font=dict(color="#e2e8f0")))
                    st.plotly_chart(ft, use_container_width=True)

                st.dataframe(op_tot[["Ordem de Produção","Horas_HH:MM","Horas","Registros","Nº Pessoas","Nº Dias"]].rename(
                    columns={"Horas_HH:MM":"Total","Horas":"h (dec.)"}).style.format({"h (dec.)":"{:.2f}"}),
                    use_container_width=True, hide_index=True)

                if not op_oper.empty:
                    st.markdown('<div class="section-header">Detalhamento por OP + Operação</div>', unsafe_allow_html=True)
                    op_lista = ["Todas"] + list(op_tot["Ordem de Produção"])
                    op_sel = st.selectbox("Filtrar OP:", op_lista)
                    df_oe = op_oper if op_sel == "Todas" else op_oper[op_oper["Ordem de Produção"]==op_sel]
                    if not df_oe.empty:
                        ycol = "Descrição" if "Descrição" in df_oe.columns else "Ordem de Produção"
                        foe = px.bar(df_oe.head(25), x="Horas", y=ycol, color="Ordem de Produção",
                            orientation="h", title="Horas por Operação", text="Horas_HH:MM")
                        foe.update_layout(height=max(350,len(df_oe.head(25))*35), **PL,
                            xaxis=dict(gridcolor="#1e2d45",color="#64748b"),
                            yaxis=dict(gridcolor="#1e2d45",color="#64748b"),
                            title=dict(font=dict(color="#e2e8f0")),
                            legend=dict(font=dict(color="#94a3b8"),bgcolor="#161b27"))
                        st.plotly_chart(foe, use_container_width=True)
                    scols = [c for c in ["Ordem de Produção","Operação","Ident. Trabalho",
                                          "Descrição","Horas_HH:MM","Horas","Registros","Pessoas"] if c in df_oe.columns]
                    st.dataframe(df_oe[scols].rename(columns={"Horas_HH:MM":"Total","Horas":"h (dec.)"})
                        .style.format({"h (dec.)":"{:.2f}"}), use_container_width=True, hide_index=True)

        # ── TAB 4 ──
        with tab4:
            if len(tp_data) > 1:
                eixo = tp_data["Data"].astype(str) + " (" + tp_data["Dia_Semana"] + ")"
                cores_pts = ["#ef4444" if r["Dia_Semana"]=="Sexta" else "#3b82f6" for _,r in tp_data.iterrows()]
                fd = go.Figure()
                fd.add_trace(go.Scatter(x=eixo, y=tp_data["Disponibilidade (%)"],
                    mode="lines+markers",
                    marker=dict(color=cores_pts,size=9,line=dict(width=1.5,color="#0f1117")),
                    line=dict(color="#3b82f6",width=2.5), name="Disponibilidade"))
                fd.add_hline(y=85, line_dash="dash", line_color="#22c55e", line_width=1.5,
                             annotation_text="Meta 85%", annotation_font_color="#22c55e")
                fd.update_layout(title=dict(text="Evolução da Disponibilidade",font=dict(color="#e2e8f0")),
                    xaxis=dict(tickangle=-30,gridcolor="#1e2d45",color="#64748b"),
                    yaxis=dict(range=[0,100],gridcolor="#1e2d45",color="#64748b"),
                    height=420, **PL)
                st.caption("🔴 Vermelho = Sexta-feira")
                st.plotly_chart(fd, use_container_width=True)

                fst = go.Figure()
                for cat, cor in [("Tempo Produtivo","#22c55e"),("Paradas Planejadas","#f59e0b"),("Paradas Não Planejadas","#ef4444")]:
                    if cat in tp_data.columns:
                        fst.add_trace(go.Bar(name=cat, x=eixo, y=tp_data[cat], marker_color=cor))
                fst.update_layout(barmode="stack",
                    title=dict(text="Composição por Dia",font=dict(color="#e2e8f0")),
                    xaxis=dict(tickangle=-30,gridcolor="#1e2d45",color="#64748b"),
                    yaxis=dict(gridcolor="#1e2d45",color="#64748b",title="Horas"),
                    height=380, **PL, legend=dict(font=dict(color="#94a3b8"),bgcolor="#161b27"))
                st.plotly_chart(fst, use_container_width=True)
            else:
                st.info("ℹ️ Uma única data — gráfico de evolução requer múltiplas datas.")

            dcols_d = ["Data","Dia_Semana","Jornada_Dia","Num_Pessoas",
                       "Tempo Produtivo_HH:MM","Paradas Planejadas_HH:MM",
                       "Paradas Não Planejadas_HH:MM","Disponibilidade (%)"]
            dcols_d = [c for c in dcols_d if c in tp_data.columns]
            st.dataframe(tp_data[dcols_d].rename(columns={
                "Dia_Semana":"Dia","Jornada_Dia":"Jornada(h)","Num_Pessoas":"Pessoas",
                "Tempo Produtivo_HH:MM":"Produtivo","Paradas Planejadas_HH:MM":"Par. Plan.",
                "Paradas Não Planejadas_HH:MM":"Par. N.Plan.","Disponibilidade (%)":"Disp. (%)"
            }).style.background_gradient(subset=["Disp. (%)"],cmap="RdYlGn",vmin=0,vmax=100)
                .format({"Disp. (%)":"{:.2f}%","Jornada(h)":"{:.1f}"}),
                use_container_width=True, hide_index=True)

        # ── TAB 5 ──
        with tab5:
            sobr = detectar_sobreposicoes(df_f)
            if sobr:
                with st.expander(f"⚠️ {len(sobr)} sobreposição(ões) detectada(s) — corrigidas automaticamente"):
                    st.caption("Intervalos sobrepostos foram fundidos antes de somar as horas.")
                    st.dataframe(pd.DataFrame(sobr), use_container_width=True, hide_index=True)

            cats = ["Todos"] + sorted(df_f["Categoria"].unique())
            ca, cb = st.columns(2)
            with ca: cat_sel = st.selectbox("Filtrar categoria:", cats)
            with cb: busca = st.text_input("🔍 Descrição:", "")

            df_show = df_f if cat_sel == "Todos" else df_f[df_f["Categoria"]==cat_sel]
            if busca: df_show = df_show[df_show["Descrição"].str.contains(busca, case=False, na=False)]

            raw_cols = ["Nome","Data","Dia_Semana","Hora inicial","Hora final",
                        "Tipo de registro do diário","Referência","N° oper.",
                        "Ident. do trabalho","Descrição","Duracao_HH:MM","Duracao_Horas","Categoria"]
            raw_cols = [c for c in raw_cols if c in df_show.columns]
            st.dataframe(df_show[raw_cols].rename(columns={
                "Dia_Semana":"Dia","N° oper.":"Oper.","Ident. do trabalho":"ID Trab.",
                "Duracao_HH:MM":"Duração","Duracao_Horas":"h (dec.)"})
                .style.format({"h (dec.)":"{:.3f}"}),
                use_container_width=True, hide_index=True)
            st.caption(f"📋 {len(df_show):,} registros")

        # ── Export ──
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("### 📥 Exportar")
        ce1, ce2 = st.columns([1, 3])
        with ce1:
            buf = gerar_excel(tp_pessoa, tp_data, paradas, op_tot, op_oper, horas_padrao)
            st.download_button("⬇️ Baixar Excel (.xlsx)", buf,
                "dashboard_disponibilidade.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with ce2:
            st.caption("Abas: Resumo · Por Data · Paradas · Horas por OP · OP + Operações")

    except Exception as e:
        st.error(f"❌ Erro: {e}"); st.exception(e)

else:
    st.markdown("""
    <div style="text-align:center;padding:60px 20px;color:#475569;">
      <div style="font-size:3.5rem;margin-bottom:16px;">📂</div>
      <h2 style="color:#64748b;font-weight:600;">Carregue a planilha para começar</h2>
      <p style="max-width:480px;margin:0 auto;font-size:.9rem;line-height:1.7;">
        Upload do <code>.xlsx</code> exportado do Dynamics 365 F&amp;O no menu lateral.<br>
        Configure a jornada e a redução de horas da sexta-feira.
      </p>
    </div>""", unsafe_allow_html=True)

    for col, icon, title, desc in zip(st.columns(3),
        ["📊","🏭","📅"],
        ["Disponibilidade","Ordens de Produção","Análise Temporal"],
        ["Por pessoa · Meta 85% · TP÷(TP+PNP)","Horas por OP e por Operação","Evolução diária · Sexta identificada"]):
        col.markdown(f"""
        <div style="background:#161b27;border:1px solid #1e2d45;border-radius:12px;padding:20px;text-align:center;">
          <div style="font-size:2rem;">{icon}</div>
          <div style="font-weight:600;color:#e2e8f0;margin:8px 0 4px;">{title}</div>
          <div style="font-size:.8rem;color:#475569;">{desc}</div>
        </div>""", unsafe_allow_html=True)